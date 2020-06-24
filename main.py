import psycopg2
import configparser
from openpyxl import load_workbook
import sys

file=['02_03_03_Matematicheskoe_obespechenie_i_administrirovanie_informatsionnykh_sistem_Razrabotka_i_administrirovanie_informatsionnykh.xlsm',
      '09_03_01_Informatika_i_vychislitelnaya_tekhnika_Avtomatizirovannye_sistemy_obrabotki_informatsii_i_upravlenia.xlsm',
      '09_03_04_Programmnaya_inzheneria_Razrabotka_programmno-informatsionnykh_sistem.xlsm',
      '09_04_01_Informatika_i_vychislitelnaya_tekhnika_Upravlenie_razrabotkoy_i_vnedreniem_IT-resheniy.xlsm']
print('Выберите файл:')
for i in range(1, 5):
    print(str(i)+')', file[i-1])
a = int(input())
if 1 <= a <= 4:
    wb = load_workbook(file[a-1])
    print(file[a-1])
else:
    print('Error')
    sys.exit()

print('Выберите ID учебного плана:')
ed_plan = int(input())

'''   Перевод недель в дни   '''
def weeks_to_days(s):
    f = s.split()
    if len(f) == 1:
        return int(f[0])*6
    f1 = f[1].split('/')
    return int(f[0])*int(f1[1])+int(f1[0])

Info = wb['Сводные данные по бюджету време']
week_days = list()
sess_days = list()
for i in range(9, 13):
    if str(Info.cell(row=i, column=2).value).lower() == 'итого':
        break
    for j in range(5, 7):
        print(str(Info.cell(row=i, column=j).value))
        week_days.append(weeks_to_days(str(Info.cell(row=i, column=j).value)))

for i in range(9, 13):
    if str(Info.cell(row=i, column=2).value).lower() == 'итого':
        break
    sess_days.append(5)
    sess_days.append(weeks_to_days(str(Info.cell(row=i, column=7).value))-5)

print(week_days)
print(sess_days)

sheet = wb['План']

disciplines = dict()
semester_disciplines = dict()
elective_blocks = dict()
rup_disciplines = dict()
rup_semesters = dict()

tmp = sheet['4:7']

for i in [0, 2, 3]:
    for j in range(0, len(tmp[i])):
        if i == 0:
            if str(tmp[i][j].value).lower() == 'индекс':
                rup_disciplines['index'] = j + 1
                disciplines['name'] = j + 2
            if str(tmp[i][j].value).lower() == 'семестр':
                rup_semesters['number'] = j + 1
            if str(tmp[i][j].value).lower() == 'код дисциплины':
                disciplines['code'] = j + 1
            if str(tmp[i][j].value).lower() == 'зачет':
                semester_disciplines['record'] = j + 1
            if str(tmp[i][j].value).lower() == 'экзамен':
                semester_disciplines['exam'] = j + 1
            if str(tmp[i][j].value).lower() == 'задания':
                semester_disciplines['task'] = j + 1
            if str(tmp[i][j].value).lower() == 'практика':
                semester_disciplines['practice_weeks_count'] = j + 1
        if i == 2:
            if str(tmp[i][j].value).lower() == 'лекции':
                semester_disciplines['lectures_count'] = j + 1
            if str(tmp[i][j].value).lower() == 'лаб. раб.':
                semester_disciplines['lab_count'] = j + 1
            if str(tmp[i][j].value).lower() == 'практ. зан.':
                semester_disciplines['practice_count'] = j + 1
            if str(tmp[i][j].value).lower() == 'всего':
                semester_disciplines['total_hours'] = j + 1
            if str(tmp[i][j].value).lower() == 'срс':
                semester_disciplines['srs_hours'] = j + 1
            if str(tmp[i][j].value).lower() == 'промежут. контроль':
                semester_disciplines['pc_hours'] = j + 1
        if i == 3:
            if str(tmp[i][j].value).lower() == 'конс.':
                semester_disciplines['consultations_hours'] = j + 1


print(sheet.cell(row=4, column=disciplines['code']).value)
if len(disciplines) != 2 or len(rup_disciplines) != 1 or len(rup_semesters) != 1 or len(semester_disciplines) != 11:
    print('Неудача')
    print(disciplines)
    print(rup_disciplines)
    print(rup_semesters)
    print(semester_disciplines)
    sys.exit()
else:
    print(disciplines)
    print(rup_semesters)
    print(rup_disciplines)

'''   Заполняю 0 пустые строки  '''
def replace_none(tmp):
    if tmp == 'None':
        return '0'
    else:
        return tmp

# Добавление лисциплины
def add_disciplines(cursor,
                    code,  # Код дисциплины         Наименование дисциплины
                    name):
    cursor.execute("INSERT INTO disciplines (code, name) VALUES (%s, %s)", (code, name))
    conn.commit()


# Добавление семестра дисциплины
def add_semester_disciplines(cursor,
                             rup_discipline,  # ID дисциплины РУП
                             rup_semester,  # ID семестра РУП
                             total_hours,  # Всего часов в семестре
                             lectures_count,  # Количество лекций в неделю
                             lab_count,  # Количество лаб в неделю
                             practice_count,  # Количество практик в неделю
                             consultations_hours,  # Часов консультаций
                             srs_hours,  # Часов СРС
                             pc_hours,  # Часов ПК
                             session_consultations_hours,  # Часов конс на сессии
                             record,  # Зачет
                             exam,  # Экзамен
                             task,  # Задание        Недель практики
                             practice_weeks_count):

    total_hours = replace_none(total_hours)
    lectures_count = replace_none(lectures_count)
    lab_count = replace_none(lab_count)
    practice_count = replace_none(practice_count)
    consultations_hours = replace_none(consultations_hours)
    srs_hours = replace_none(srs_hours)
    pc_hours = replace_none(pc_hours)
    session_consultations_hours = replace_none(session_consultations_hours)
    record = replace_none(record)
    exam = replace_none(exam)
    task = replace_none(task)
    practice_weeks_count = replace_none(practice_weeks_count)

    cursor.execute("INSERT INTO semester_disciplines (rup_discipline,"
                   "rup_semester,"
                   "total_hours,"
                   "lectures_count,"
                   "lab_count,"
                   "practice_count,"
                   "consultations_hours,"
                   "srs_hours,"
                   "pc_hours,"
                   "session_consultations_hours,"
                   "record,"
                   "exam,"
                   "task,"
                   "practice_weeks_count) "
                   "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                   (rup_discipline,
                    rup_semester,
                    total_hours,
                    lectures_count,
                    lab_count,
                    practice_count,
                    consultations_hours,
                    srs_hours,
                    pc_hours,
                    session_consultations_hours,
                    record,
                    exam,
                    task,
                    practice_weeks_count))
    conn.commit()


# Добавление блока элективов
def add_elective_blocks(cursor,
                        number,  # Номер блока элективов          ID рабочего учебного плана
                        educational_plan=ed_plan):
    cursor.execute("INSERT INTO elective_blocks (number, educational_plan) VALUES (%s, %s)", (number, educational_plan))
    conn.commit()


# Добавление дисциплины РУП
def add_rup_disciplines(cursor,
                        index,  # Индекс дисциплины в РУП
                        discipline,  # ID дисциплины
                        educational_plan=ed_plan,
                        practice_type=None):
    cursor.execute("INSERT INTO rup_disciplines (index, discipline, educational_plan, practice_type) VALUES (%s, %s, %s, %s)",
                   (index, discipline, educational_plan, practice_type))
    conn.commit()


# Добавление семестра РУП
def add_rup_semesters(cursor,
                      number,  # Номер семестра РУП
                      days_in_a_session,  # Дней на сессию
                      days_in_a_week,  # Дней на зачетной неделе             ID рабочего учебного плана
                      educational_plan=ed_plan):
    cursor.execute(
        "INSERT INTO rup_semesters (number, days_in_a_session, days_in_a_week, educational_plan) VALUES (%s, %s, %s, %s)",
        (number, days_in_a_session, days_in_a_week, educational_plan,))
    conn.commit()

# Добавление типа практики
def add_practice_types(cursor,
                      practice_kind,
                      name,
                      opop_component=2):
    cursor.execute(
        "INSERT INTO practice_types (practice_kind, name, opop_component) VALUES (%s, %s, %s)",
        (practice_kind, name, opop_component,))
    conn.commit()

config = configparser.ConfigParser()  # создаём объекта парсера
config.read("settings.ini")  # читаем конфиг

conn = psycopg2.connect(dbname=config['Postgres']['dbname'], user=config['Postgres']['user'],
                        password=config['Postgres']['password'], host=config['Postgres']['host'])
cursor = conn.cursor()

'''   Заполнение дисциплин (готово)   '''
rowNum = 9
while str(sheet.cell(row=rowNum, column=2).value) != 'Б2':
    if str(sheet.cell(row=rowNum, column=4).value) == 'None':
        rowNum = rowNum + 1
        continue

# 1) Дисциплины
    cursor.execute('SELECT id FROM disciplines where code like %s',
                   (str(sheet.cell(row=rowNum, column=disciplines['code']).value),))
    record = cursor.fetchone()
    if record is None:
        add_disciplines(cursor, str(sheet.cell(row=rowNum, column=disciplines['code']).value),
                        str(sheet.cell(row=rowNum, column=disciplines['name']).value))

# 2) Дисциплины РУП
    cursor.execute('SELECT id FROM disciplines where code like %s',
                  (str(sheet.cell(row=rowNum, column=disciplines['code']).value),))
    record = cursor.fetchone()
    # print(record[0])
    add_rup_disciplines(cursor, str(sheet.cell(row=rowNum, column=rup_disciplines['index']).value), record[0])

# 3) Семестры РУП
    print(str(sheet.cell(row=rowNum, column=rup_semesters['number']).value))
    cursor.execute('SELECT id FROM rup_semesters where number = %s and educational_plan = %s',
                   (str(sheet.cell(row=rowNum, column=rup_semesters['number']).value), ed_plan,))
    record = cursor.fetchone()
    if record is None:
        print(int(sheet.cell(row=rowNum, column=rup_semesters['number']).value)-1)
        add_rup_semesters(cursor, str(sheet.cell(row=rowNum, column=rup_semesters['number']).value),
                          sess_days[int(sheet.cell(row=rowNum, column=rup_semesters['number']).value)-1],
                          week_days[int(sheet.cell(row=rowNum, column=rup_semesters['number']).value)-1])

# 4) Семестры дисциплин
    cursor.execute('SELECT id FROM rup_disciplines where index like %s',
                   (str(sheet.cell(row=rowNum, column=rup_disciplines['index']).value),))
    discipl = cursor.fetchone()
    # print(discipl[0])
    cursor.execute('SELECT id FROM rup_semesters where number = %s',
                   (str(sheet.cell(row=rowNum, column=rup_semesters['number']).value),))
    sem = cursor.fetchone()
    # print(sem[0])
    cursor.execute('SELECT id FROM semester_disciplines where rup_discipline = %s and rup_semester=%s',
                   (discipl[0], sem[0],))
    exist = cursor.fetchone()
    if exist is None:
        add_semester_disciplines(cursor,
                                 discipl[0],  # ID дисциплины РУП
                                 sem[0],  # ID семестра РУП
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['total_hours']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['lectures_count']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['lab_count']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['practice_count']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['consultations_hours']).value),
                                 # Часов консультаций
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['srs_hours']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['pc_hours']).value),
                                 0,  # Часов конс на сессии
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['record']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['exam']).value),
                                 str(sheet.cell(row=rowNum, column=semester_disciplines['task']).value),
                                 0)  # str(sheet.cell(row=rowNum, column=semester_disciplines['practice_weeks_count']).value)

    rowNum = rowNum + 1



'''   Практика   '''
# while str(sheet.cell(row=rowNum, column=2).value) != 'Б3':
#     if str(sheet.cell(row=rowNum, column=2).value) == 'None':
#         rowNum = rowNum + 1
#         continue
#     if str(sheet.cell(row=rowNum, column=3).value) == 'Учебная практика':
#         cursor.execute('SELECT id FROM practice_kinds where name like %s',
#                        (str(sheet.cell(row=rowNum, column=3).value),))
#         practice_types = cursor.fetchone()
#         rowNum = rowNum + 1
#         continue
#     elif str(sheet.cell(row=rowNum, column=3).value) == 'Производственная практика':
#         cursor.execute('SELECT id FROM practice_kinds where name like %s',
#                        (str(sheet.cell(row=rowNum, column=3).value),))
#         practice_types = cursor.fetchone()
#         rowNum = rowNum + 1
#         continue
#
# # 1) Тип практики
#     cursor.execute('SELECT id FROM practice_types where practice_kind = %s and name like %s',
#                    (practice_types[0], str(sheet.cell(row=rowNum, column=3).value),))
#     record = cursor.fetchone()
#     if record is None:
#         add_practice_types(cursor, practice_types[0], str(sheet.cell(row=rowNum, column=3).value))
#
# # 2) Дисциплины РУП
#     cursor.execute('SELECT id FROM disciplines where code like %s',
#                   (str(sheet.cell(row=rowNum, column=disciplines['code']).value),))
#     record = cursor.fetchone()
#     print(record[0])
#     cursor.execute('SELECT id FROM practice_types where practice_kind = %s and name like %s',
#                    (practice_types[0], str(sheet.cell(row=rowNum, column=3).value),))
#     practice_types = cursor.fetchone()
#     add_rup_disciplines(cursor, str(sheet.cell(row=rowNum, column=rup_disciplines['index']).value), record[0],
#                         practice_type=practice_types)
#
# # 4) Семестры дисциплин
#     cursor.execute('SELECT id FROM rup_disciplines where index like %s',
#                    (str(sheet.cell(row=rowNum, column=rup_disciplines['index']).value),))
#     discipl = cursor.fetchone()
#     print(discipl[0])
#     cursor.execute('SELECT id FROM rup_semesters where number = %s',
#                    (str(sheet.cell(row=rowNum, column=rup_semesters['number']).value),))
#     sem = cursor.fetchone()
#     print(sem[0])
#
#     add_semester_disciplines(cursor,
#                              discipl[0],  # ID дисциплины РУП
#                              sem[0],  # ID семестра РУП
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['total_hours']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['lectures_count']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['lab_count']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['practice_count']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['consultations_hours']).value),
#                              # Часов консультаций
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['srs_hours']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['pc_hours']).value),
#                              0,  # Часов конс на сессии
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['record']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['exam']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['task']).value),
#                              str(sheet.cell(row=rowNum, column=semester_disciplines['practice_weeks_count']).value))  # str(sheet.cell(row=rowNum, column=semester_disciplines['practice_weeks_count']).value)
#
#     rowNum = rowNum + 1
#

cursor.close()
conn.close()
